from openpyxl import Workbook, load_workbook


def create_sheet(size):
    book = Workbook()

    for i in range(size):
        nme = "Chromosome" + str(i)
        book.create_sheet(nme)
        nme = book.active
        nme.cell(row=5, column=5).value = 'Hi'
    book.save('test.xlsx')


def insert_into_excel(row, column, data):
    book = Workbook()
    sheet = book.active
    sheet.cell(row=row, column=column).value = data
    book.save('test.xlsx')


if __name__ == "__main__":
    # row, column, data = 5, 5, 'CE 587'
    # insert_into_excel(row=row, column=column, data=data)
    c = [5, 8, 9]
    for chromosome in c:
        print(c.index(chromosome))
